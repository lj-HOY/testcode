import pandas as pd
from openpyxl import load_workbook


# Result = load_workbook('统计分析.xlsx')

# Result_sheet = Result['Sheet1']
data = pd.read_excel('预测结果分析_total2.xlsx')  # 导入参赛信息
b = data['不好的图序'].value_counts() #ocean_proximity为属性标签
g = data['好的图序'].value_counts() #ocean_proximity为属性标签.
n = data['一般的图序'].value_counts() #ocean_proximity为属性标签

# print(len(b))
# print(b)
# bb=b.tolist()
# print(bb)
# for i in range(1, len(b)+1):#记录不好的
#     row_w = Result_sheet.max_row + 1
#     for j in range(1, 2):
#         Result_sheet.cell(row_w, column=1).value = b[i,j]
#
#

print(b)
print(g)
print(n)
print(abc)